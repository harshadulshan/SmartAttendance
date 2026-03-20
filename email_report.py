import smtplib
import pandas as pd
import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ─────────────────────────────────────────
# EMAIL CONFIGURATION
# ─────────────────────────────────────────
EMAIL_CONFIG = {
    'sender_email'    : '',    # ← add Your Gmail
    'sender_password' : '',        # ← you Gmail App Password
    'receiver_email'  : '',       # ← add receives gamil
    'smtp_server'     : 'smtp.gmail.com',
    'smtp_port'       : 587
}

def load_today_attendance():
    """Load today's attendance data."""
    today = datetime.now().strftime('%Y-%m-%d')
    file  = f'attendance/attendance_{today}.csv'
    if os.path.exists(file):
        df = pd.read_csv(file)
        # Add Arrival column if missing
        if 'Arrival' not in df.columns:
            df['Arrival'] = 'On Time 🟢'
        return df, today
    return pd.DataFrame(columns=[
        'Name', 'Date', 'Time', 'Status', 'Arrival'
    ]), today

def create_excel_report(df, today):
    """Create formatted Excel report."""
    path = f'attendance/daily_report_{today}.xlsx'
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = 'Daily Report'

    # Title
    ws.merge_cells('A1:E1')
    title_cell       = ws['A1']
    title_cell.value = f'Smart Attendance Report — {today}'
    title_cell.font  = Font(bold=True, size=14, color='00D9FF')
    title_cell.fill  = PatternFill(start_color='0D1117',
                                   end_color='0D1117',
                                   fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center')

    # Headers
    headers     = ['Name', 'Date', 'Time', 'Status', 'Arrival']
    header_fill = PatternFill(start_color='161B22',
                              end_color='161B22',
                              fill_type='solid')
    header_font = Font(color='00D9FF', bold=True)

    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row, record in enumerate(df.to_dict('records'), 3):
        for col, key in enumerate(headers, 1):
            cell       = ws.cell(row=row, column=col,
                                value=record.get(key, ''))
            cell.alignment = Alignment(horizontal='center')
            if 'Late' in str(record.get('Arrival', '')):
                cell.fill = PatternFill(start_color='2D1B00',
                                       end_color='2D1B00',
                                       fill_type='solid')
                cell.font = Font(color='F59E0B')

    # Column widths
        # Column widths
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                try:
                    if cell.column_letter and cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                        col_letter = cell.column_letter
                except AttributeError:
                    continue
            if col_letter:
                ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(path)
    return path

def create_html_report(df, today):
    """Create beautiful HTML email body."""
    total     = len(df)
    on_time   = len(df[~df['Arrival'].str.contains('Late', na=False)]) if not df.empty else 0
    late      = len(df[df['Arrival'].str.contains('Late', na=False)])  if not df.empty else 0

    # Build table rows
    rows = ""
    if not df.empty:
        for _, row in df.iterrows():
            late_style = "color:#F59E0B;background:#2D1B00;" if "Late" in str(row.get('Arrival','')) else ""
            rows += f"""
            <tr style="{late_style}">
                <td style="padding:10px;border:1px solid #30363d">{row.get('Name','')}</td>
                <td style="padding:10px;border:1px solid #30363d">{row.get('Time','')}</td>
                <td style="padding:10px;border:1px solid #30363d">{row.get('Status','')}</td>
                <td style="padding:10px;border:1px solid #30363d">{row.get('Arrival','')}</td>
            </tr>"""
    else:
        rows = '<tr><td colspan="4" style="padding:20px;text-align:center">No attendance recorded today</td></tr>'

    html = f"""
    <html>
    <body style="background:#0D1117;color:#fff;font-family:Segoe UI,sans-serif;padding:30px">

        <div style="max-width:700px;margin:0 auto">

            <!-- Header -->
            <div style="background:#161B22;padding:25px;border-radius:12px;
                        border:1px solid #30363d;margin-bottom:20px">
                <h1 style="color:#00D9FF;margin:0">🎯 Smart Attendance Report</h1>
                <p style="color:#aaa;margin:5px 0 0 0">{today}</p>
            </div>

            <!-- Stats -->
            <div style="display:flex;gap:15px;margin-bottom:20px">
                <div style="background:#161B22;padding:20px;border-radius:12px;
                            border:1px solid #30363d;flex:1;text-align:center">
                    <h2 style="color:#00D9FF;margin:0;font-size:2.5rem">{total}</h2>
                    <p style="color:#aaa;margin:5px 0 0 0">Total Present</p>
                </div>
                <div style="background:#161B22;padding:20px;border-radius:12px;
                            border:1px solid #30363d;flex:1;text-align:center">
                    <h2 style="color:#22C55E;margin:0;font-size:2.5rem">{on_time}</h2>
                    <p style="color:#aaa;margin:5px 0 0 0">On Time 🟢</p>
                </div>
                <div style="background:#161B22;padding:20px;border-radius:12px;
                            border:1px solid #30363d;flex:1;text-align:center">
                    <h2 style="color:#F59E0B;margin:0;font-size:2.5rem">{late}</h2>
                    <p style="color:#aaa;margin:5px 0 0 0">Late 🔴</p>
                </div>
            </div>

            <!-- Table -->
            <div style="background:#161B22;padding:20px;border-radius:12px;
                        border:1px solid #30363d">
                <h3 style="color:#fff;margin:0 0 15px 0">📋 Attendance Log</h3>
                <table style="width:100%;border-collapse:collapse">
                    <thead>
                        <tr style="background:#0D1117">
                            <th style="padding:10px;border:1px solid #30363d;
                                      color:#00D9FF;text-align:left">Name</th>
                            <th style="padding:10px;border:1px solid #30363d;
                                      color:#00D9FF;text-align:left">Time</th>
                            <th style="padding:10px;border:1px solid #30363d;
                                      color:#00D9FF;text-align:left">Status</th>
                            <th style="padding:10px;border:1px solid #30363d;
                                      color:#00D9FF;text-align:left">Arrival</th>
                        </tr>
                    </thead>
                    <tbody>{rows}</tbody>
                </table>
            </div>

            <!-- Footer -->
            <p style="color:#555;text-align:center;margin-top:20px;font-size:0.8rem">
                Generated by Smart Attendance System — {datetime.now().strftime('%H:%M:%S')}
            </p>
        </div>
    </body>
    </html>
    """
    return html

def send_report():
    """Send daily attendance report via email."""
    print("\n📧 Preparing email report...")

    df, today = load_today_attendance()

    # Create Excel attachment
    excel_path = create_excel_report(df, today)
    html_body  = create_html_report(df, today)

    # Build email
    msg                    = MIMEMultipart('alternative')
    msg['Subject']         = f"📊 Attendance Report — {today}"
    msg['From']            = EMAIL_CONFIG['sender_email']
    msg['To']              = EMAIL_CONFIG['receiver_email']

    # HTML body
    msg.attach(MIMEText(html_body, 'html'))

    # Excel attachment
    with open(excel_path, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                       f'attachment; filename=attendance_{today}.xlsx')
        msg.attach(part)

    # Send
    try:
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'],
                         EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'],
                        EMAIL_CONFIG['sender_password'])
            server.send_message(msg)
        print(f"✅ Report sent to {EMAIL_CONFIG['receiver_email']}!")
        return True
    except Exception as e:
        print(f"❌ Email failed: {e}")
        print("💡 Make sure you set up Gmail App Password correctly!")
        return False

if __name__ == '__main__':
    print("=" * 45)
    print("  📧 AUTO EMAIL REPORT SYSTEM")
    print("=" * 45)
    send_report()