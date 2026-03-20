import dash
from dash import dcc, html, dash_table
from dash.dependencies import Input, Output, State
import plotly.express as px
import plotly.graph_objects as go
import pandas as pd
import os
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import dash_auth

# ─────────────────────────────────────────
# DATA LOADING
# ─────────────────────────────────────────
def load_all_attendance():
    all_data = []
    if not os.path.exists('attendance'):
        os.makedirs('attendance')
        return pd.DataFrame(columns=[
            'Name', 'Date', 'Time', 'Status', 'Arrival'
        ])
    for file in os.listdir('attendance'):
        if file.startswith('attendance_') and file.endswith('.csv'):
            df = pd.read_csv(f'attendance/{file}')
            if 'Arrival' not in df.columns:
                df['Arrival'] = 'On Time 🟢'
            all_data.append(df)
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return pd.DataFrame(columns=[
        'Name', 'Date', 'Time', 'Status', 'Arrival'
    ])

def load_unknown_log():
    all_data = []
    if os.path.exists('attendance'):
        for file in os.listdir('attendance'):
            if file.startswith('unknown_log_') and file.endswith('.csv'):
                df = pd.read_csv(f'attendance/{file}')
                all_data.append(df)
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return pd.DataFrame(columns=['Date', 'Time', 'Snapshot'])

def export_to_excel(df, period='daily'):
    path = f'attendance/report_{period}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = f'{period.capitalize()} Report'

    header_fill = PatternFill(start_color='0D1117',
                              end_color='0D1117', fill_type='solid')
    header_font = Font(color='00D9FF', bold=True, size=12)

    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, record in enumerate(df.to_dict('records'), 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=record.get(key, ''))

    for col in ws.columns:
        max_len    = 0
        col_letter = None
        for cell in col:
            try:
                if cell.column_letter and cell.value:
                    max_len    = max(max_len, len(str(cell.value)))
                    col_letter = cell.column_letter
            except AttributeError:
                continue
        if col_letter:
            ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(path)
    return path

# ─────────────────────────────────────────
# APP + LOGIN
# ─────────────────────────────────────────
app       = dash.Dash(__name__, suppress_callback_exceptions=True)
app.title = "Smart Attendance Dashboard"

VALID_USERS = {
    'admin'  : '',

}
auth = dash_auth.BasicAuth(app, VALID_USERS)

# ─────────────────────────────────────────
# THEMES
# ─────────────────────────────────────────
DARK_THEME = {
    'bg'      : '#0D1117',
    'card'    : '#161B22',
    'border'  : '#30363d',
    'text'    : '#ffffff',
    'subtext' : '#aaaaaa',
    'accent1' : '#00D9FF',
    'accent2' : '#A855F7',
    'accent3' : '#22C55E',
    'accent4' : '#F59E0B',
    'danger'  : '#EF4444',
}

LIGHT_THEME = {
    'bg'      : '#F6F8FA',
    'card'    : '#FFFFFF',
    'border'  : '#D0D7DE',
    'text'    : '#1F2328',
    'subtext' : '#656D76',
    'accent1' : '#0969DA',
    'accent2' : '#8250DF',
    'accent3' : '#1A7F37',
    'accent4' : '#9A6700',
    'danger'  : '#CF222E',
}

def get_theme(theme):
    return DARK_THEME if theme == 'dark' else LIGHT_THEME

def chart_style(t):
    return {
        'paper_bgcolor' : t['card'],
        'plot_bgcolor'  : t['card'],
        'font'          : {'color': t['text']},
        'margin'        : dict(t=30, b=10, l=10, r=10),
        'xaxis'         : {'gridcolor': t['border']},
        'yaxis'         : {'gridcolor': t['border']}
    }

def stat_card(value, label, color, theme):
    t = get_theme(theme)
    return html.Div([
        html.H2(value, style={
            'color': color, 'fontSize': '3rem', 'margin': '0'
        }),
        html.P(label, style={
            'color': t['subtext'], 'margin': '5px 0 0 0'
        })
    ], style={
        'background'   : t['card'],
        'padding'      : '25px',
        'borderRadius' : '12px',
        'textAlign'    : 'center',
        'border'       : f"1px solid {t['border']}",
        'flex'         : '1'
    })

# ─────────────────────────────────────────
# LAYOUT
# ─────────────────────────────────────────
app.layout = html.Div(id='main-container', children=[

    dcc.Interval(id='interval', interval=5000, n_intervals=0),
    dcc.Store(id='theme-store', data='dark'),
    dcc.Download(id='download-excel'),

    # Header
    html.Div(id='header', children=[
        html.Div([
            html.H1("🎯 Smart Attendance System",
                    id='header-title',
                    style={'margin': '0', 'fontSize': '2rem'}),
            html.P(id='header-date',
                   style={'margin': '5px 0 0 0', 'fontSize': '0.9rem'}),
        ]),
        html.Div([
            html.Button("🌙 Dark", id='theme-btn', style={
                'padding'     : '8px 20px',
                'borderRadius': '20px',
                'border'      : '1px solid #30363d',
                'cursor'      : 'pointer',
                'fontWeight'  : 'bold',
                'marginRight' : '10px'
            }),
            html.Button("📥 Export Excel", id='export-btn', style={
                'padding'     : '8px 20px',
                'borderRadius': '20px',
                'border'      : 'none',
                'cursor'      : 'pointer',
                'fontWeight'  : 'bold',
                'background'  : '#00D9FF',
                'color'       : '#000'
            }),
        ], style={'display': 'flex', 'alignItems': 'center'}),
    ], style={
        'display'        : 'flex',
        'justifyContent' : 'space-between',
        'alignItems'     : 'center',
        'padding'        : '20px 40px',
        'borderBottom'   : '2px solid #30363d'
    }),

    # Content
    html.Div(id='content', children=[
        dcc.Tabs(id='tabs', value='today', children=[
            dcc.Tab(
                label='📅 Today',
                value='today',
                style={'color': '#aaaaaa', 'backgroundColor': '#0D1117'},
                selected_style={'color': '#00D9FF', 'backgroundColor': '#161B22',
                                'borderTop': '3px solid #00D9FF'}
            ),
            dcc.Tab(
                label='📆 Weekly',
                value='weekly',
                style={'color': '#aaaaaa', 'backgroundColor': '#0D1117'},
                selected_style={'color': '#00D9FF', 'backgroundColor': '#161B22',
                                'borderTop': '3px solid #00D9FF'}
            ),
            dcc.Tab(
                label='🗓️ Monthly',
                value='monthly',
                style={'color': '#aaaaaa', 'backgroundColor': '#0D1117'},
                selected_style={'color': '#00D9FF', 'backgroundColor': '#161B22',
                                'borderTop': '3px solid #00D9FF'}
            ),
            dcc.Tab(
                label='⚠️ Unknown',
                value='unknown',
                style={'color': '#aaaaaa', 'backgroundColor': '#0D1117'},
                selected_style={'color': '#00D9FF', 'backgroundColor': '#161B22',
                                'borderTop': '3px solid #00D9FF'}
            ),
        ], style={'marginBottom': '25px'}),
        html.Div(id='tab-content'),
    ], style={'padding': '25px 40px'}),

], style={'minHeight': '100vh', 'fontFamily': 'Segoe UI, sans-serif'})

# ─────────────────────────────────────────
# THEME CALLBACK
# ─────────────────────────────────────────
@app.callback(
    [Output('theme-store',    'data'),
     Output('theme-btn',      'children'),
     Output('main-container', 'style'),
     Output('header',         'style'),
     Output('header-title',   'style'),
     Output('header-date',    'style'),
     Output('theme-btn',      'style'),
     Output('content',        'style')],
    [Input('theme-btn', 'n_clicks')],
    [State('theme-store', 'data')]
)
def toggle_theme(n, current_theme):
    if n is None:
        theme = 'dark'
    else:
        theme = 'light' if current_theme == 'dark' else 'dark'
    t     = get_theme(theme)
    label = "☀️ Light" if theme == 'dark' else "🌙 Dark"
    return (
        theme, label,
        {'minHeight': '100vh', 'fontFamily': 'Segoe UI, sans-serif',
         'background': t['bg']},
        {'display': 'flex', 'justifyContent': 'space-between',
         'alignItems': 'center', 'padding': '20px 40px',
         'borderBottom': f"2px solid {t['border']}",
         'background': t['card']},
        {'margin': '0', 'fontSize': '2rem', 'color': t['accent1']},
        {'margin': '5px 0 0 0', 'fontSize': '0.9rem', 'color': t['subtext']},
        {'padding': '8px 20px', 'borderRadius': '20px',
         'border': f"1px solid {t['border']}", 'cursor': 'pointer',
         'fontWeight': 'bold', 'background': t['card'],
         'color': t['text'], 'marginRight': '10px'},
        {'padding': '25px 40px', 'background': t['bg']}
    )

# ─────────────────────────────────────────
# HEADER DATE
# ─────────────────────────────────────────
@app.callback(
    Output('header-date', 'children'),
    Input('interval', 'n_intervals')
)
def update_date(n):
    return datetime.now().strftime('%A, %B %d %Y  |  %H:%M:%S')

# ─────────────────────────────────────────
# TAB CONTENT
# ─────────────────────────────────────────
@app.callback(
    Output('tab-content', 'children'),
    [Input('tabs', 'value'),
     Input('interval', 'n_intervals')],
    [State('theme-store', 'data')]
)
def render_tab(tab, n, theme):
    t     = get_theme(theme)
    df    = load_all_attendance()
    today = datetime.now().strftime('%Y-%m-%d')

    card_style = {
        'background'   : t['card'],
        'padding'      : '20px',
        'borderRadius' : '12px',
        'border'       : f"1px solid {t['border']}",
        'marginBottom' : '20px'
    }

    table_style = {
        'style_table' : {'overflowX': 'auto'},
        'style_cell'  : {
            'backgroundColor' : t['bg'],
            'color'           : t['text'],
            'border'          : f"1px solid {t['border']}",
            'padding'         : '12px',
            'textAlign'       : 'left'
        },
        'style_header' : {
            'backgroundColor' : t['card'],
            'color'           : t['accent1'],
            'fontWeight'      : 'bold',
            'border'          : f"1px solid {t['border']}"
        }
    }

    # ════════════════════════════════════
    # TODAY TAB
    # ════════════════════════════════════
    if tab == 'today':
        today_df         = df[df['Date'] == today] if not df.empty else pd.DataFrame()
        total_registered = df['Name'].nunique() if not df.empty else 0
        total_present    = today_df['Name'].nunique() if not today_df.empty else 0
        late_count       = len(today_df[today_df['Arrival'].str.contains('Late', na=False)]) if not today_df.empty else 0
        on_time_count    = total_present - late_count
        rate             = f"{int((total_present/total_registered)*100)}%" if total_registered > 0 else "0%"

        absent  = max(0, total_registered - total_present)
        pie_fig = go.Figure(data=[go.Pie(
            labels=['On Time', 'Absent', 'Late'],
            values=[on_time_count, absent, late_count],
            marker_colors=[t['accent3'], t['danger'], t['accent4']],
            hole=0.45
        )])
        pie_fig.update_layout(**chart_style(t))

        if not today_df.empty:
            today_df = today_df.copy()
            today_df['Hour'] = pd.to_datetime(
                today_df['Time'], format='%H:%M:%S'
            ).dt.hour
            hourly   = today_df.groupby('Hour').size().reset_index(name='Count')
            time_fig = px.bar(hourly, x='Hour', y='Count',
                             color_discrete_sequence=[t['accent1']],
                             title='Arrival Timeline')
        else:
            time_fig = go.Figure()

        time_fig.update_layout(**chart_style(t))

        return html.Div([
            html.Div([
                stat_card(str(total_present), "Present Today",    t['accent1'], theme),
                stat_card(str(late_count),    "Late Arrivals 🔴", t['accent4'], theme),
                stat_card(str(on_time_count), "On Time 🟢",       t['accent3'], theme),
                stat_card(rate,               "Attendance Rate",  t['accent2'], theme),
            ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '20px'}),

            html.Div([
                html.Div([
                    dcc.Graph(figure=time_fig, style={'height': '280px'})
                ], style={**card_style, 'flex': '2'}),
                html.Div([
                    dcc.Graph(figure=pie_fig, style={'height': '280px'})
                ], style={**card_style, 'flex': '1'}),
            ], style={'display': 'flex', 'gap': '20px'}),

            html.Div([
                html.H3("📋 Today's Attendance Log",
                        style={'color': t['text'], 'marginBottom': '15px'}),
                dash_table.DataTable(
                    data=today_df.to_dict('records') if not today_df.empty else [],
                    columns=[{'name': c, 'id': c} for c in today_df.columns] if not today_df.empty else [],
                    filter_action='native',
                    sort_action='native',
                    page_size=10,
                    **table_style
                )
            ], style=card_style),
        ])

    # ════════════════════════════════════
    # WEEKLY TAB
    # ════════════════════════════════════
    elif tab == 'weekly':
        week_start = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        week_df    = df[df['Date'] >= week_start] if not df.empty else pd.DataFrame()

        if not week_df.empty:
            daily         = week_df.groupby('Date')['Name'].nunique().reset_index()
            daily.columns = ['Date', 'Count']
            bar_fig       = px.bar(daily, x='Date', y='Count',
                                  color_discrete_sequence=[t['accent1']],
                                  title='Daily Attendance This Week')

            person         = week_df.groupby('Name')['Date'].nunique().reset_index()
            person.columns = ['Name', 'Days']
            line_fig       = px.bar(person, x='Name', y='Days',
                                   color='Days',
                                   color_continuous_scale=[t['accent2'], t['accent1']],
                                   title='Individual Attendance This Week')

            late_df  = week_df[week_df['Arrival'].str.contains('Late', na=False)]
            late_grp = late_df.groupby('Name').size().reset_index(name='Late Count')
            late_fig = px.bar(late_grp, x='Name', y='Late Count',
                             color_discrete_sequence=[t['accent4']],
                             title='Late Arrivals This Week')
        else:
            bar_fig  = go.Figure()
            line_fig = go.Figure()
            late_fig = go.Figure()

        for fig in [bar_fig, line_fig, late_fig]:
            fig.update_layout(**chart_style(t))

        total_present = week_df['Name'].nunique() if not week_df.empty else 0
        total_records = len(week_df)               if not week_df.empty else 0
        late_total    = len(week_df[week_df['Arrival'].str.contains('Late', na=False)]) if not week_df.empty else 0

        return html.Div([
            html.Div([
                stat_card(str(total_present), "Unique Attendees", t['accent1'], theme),
                stat_card(str(total_records), "Total Records",    t['accent2'], theme),
                stat_card(str(late_total),    "Late Arrivals 🔴", t['accent4'], theme),
                stat_card("7",               "Days Range",        t['accent3'], theme),
            ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '20px'}),

            html.Div([dcc.Graph(figure=bar_fig,  style={'height': '280px'})], style=card_style),
            html.Div([dcc.Graph(figure=line_fig, style={'height': '280px'})], style=card_style),
            html.Div([dcc.Graph(figure=late_fig, style={'height': '280px'})], style=card_style),
        ])

    # ════════════════════════════════════
    # MONTHLY TAB
    # ════════════════════════════════════
    elif tab == 'monthly':
        current_month = datetime.now().strftime('%Y-%m')
        month_df      = df[df['Date'].str.startswith(current_month)] if not df.empty else pd.DataFrame()

        if not month_df.empty:
            month_df = month_df.copy()

            daily         = month_df.groupby('Date')['Name'].nunique().reset_index()
            daily.columns = ['Date', 'Count']
            trend_fig     = px.line(daily, x='Date', y='Count',
                                   markers=True,
                                   color_discrete_sequence=[t['accent1']],
                                   title='Monthly Attendance Trend')

            person         = month_df.groupby('Name')['Date'].nunique().reset_index()
            person.columns = ['Name', 'Days Present']
            person_fig     = px.bar(person, x='Name', y='Days Present',
                                   color='Days Present',
                                   color_continuous_scale=['#EF4444', '#F59E0B', '#22C55E'],
                                   title='Monthly Individual Performance')

            # ── HEATMAP ──────────────────────────
            month_df['DayOfWeek'] = pd.to_datetime(
                month_df['Date']
            ).dt.day_name()
            month_df['Hour'] = pd.to_datetime(
                month_df['Time'], format='%H:%M:%S'
            ).dt.hour

            heatmap_data   = month_df.groupby(
                ['DayOfWeek', 'Hour']
            ).size().reset_index(name='Count')

            day_order = [
                'Monday', 'Tuesday', 'Wednesday',
                'Thursday', 'Friday', 'Saturday', 'Sunday'
            ]

            heatmap_pivot = heatmap_data.pivot(
                index='DayOfWeek',
                columns='Hour',
                values='Count'
            ).reindex(day_order).fillna(0)

            heatmap_fig = go.Figure(data=go.Heatmap(
                z=heatmap_pivot.values,
                x=[f'{h:02d}:00' for h in heatmap_pivot.columns],
                y=heatmap_pivot.index.tolist(),
                colorscale=[
                    [0.0, t['card']],
                    [0.3, t['accent2']],
                    [0.6, t['accent1']],
                    [1.0, t['accent3']]
                ],
                showscale=True,
                hoverongaps=False
            ))
            heatmap_fig.update_layout(
                **chart_style(t),
                title='🗓️ Attendance Heatmap — Day vs Hour'
            )

        else:
            trend_fig   = go.Figure()
            person_fig  = go.Figure()
            heatmap_fig = go.Figure()

        for fig in [trend_fig, person_fig, heatmap_fig]:
            fig.update_layout(**chart_style(t))

        total_days    = month_df['Date'].nunique()  if not month_df.empty else 0
        total_records = len(month_df)               if not month_df.empty else 0
        avg_per_day   = round(total_records / total_days, 1) if total_days > 0 else 0
        late_total    = len(month_df[month_df['Arrival'].str.contains('Late', na=False)]) if not month_df.empty else 0

        return html.Div([
            html.Div([
                stat_card(str(total_days),    "Days Recorded",   t['accent1'], theme),
                stat_card(str(total_records), "Total Records",   t['accent2'], theme),
                stat_card(str(avg_per_day),   "Avg Per Day",     t['accent3'], theme),
                stat_card(str(late_total),    "Late This Month", t['accent4'], theme),
            ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '20px'}),

            html.Div([dcc.Graph(figure=trend_fig,   style={'height': '300px'})], style=card_style),
            html.Div([dcc.Graph(figure=person_fig,  style={'height': '300px'})], style=card_style),

            # Heatmap
            html.Div([
                html.H3("🗺️ Attendance Heatmap",
                        style={'color': t['text'], 'marginBottom': '15px'}),
                dcc.Graph(figure=heatmap_fig, style={'height': '350px'})
            ], style=card_style),

            # Monthly table
            html.Div([
                html.H3("📋 Monthly Attendance Log",
                        style={'color': t['text'], 'marginBottom': '15px'}),
                dash_table.DataTable(
                    data=month_df.to_dict('records') if not month_df.empty else [],
                    columns=[{'name': c, 'id': c} for c in month_df.columns] if not month_df.empty else [],
                    filter_action='native',
                    sort_action='native',
                    page_size=15,
                    **table_style
                )
            ], style=card_style),
        ])

    # ════════════════════════════════════
    # UNKNOWN TAB
    # ════════════════════════════════════
    elif tab == 'unknown':
        unknown_df    = load_unknown_log()
        total_unknown = len(unknown_df)
        today_unknown = len(unknown_df[unknown_df['Date'] == today]) if not unknown_df.empty else 0

        if not unknown_df.empty:
            daily_unk = unknown_df.groupby('Date').size().reset_index(name='Count')
            unk_fig   = px.bar(daily_unk, x='Date', y='Count',
                              color_discrete_sequence=[t['danger']],
                              title='Unknown Face Detections Per Day')
            unk_fig.update_layout(**chart_style(t))
        else:
            unk_fig = go.Figure()
            unk_fig.update_layout(**chart_style(t))

        return html.Div([
            html.Div([
                stat_card(str(total_unknown), "Total Unknown Detected", t['danger'],  theme),
                stat_card(str(today_unknown), "Today's Unknown Faces",  t['accent4'], theme),
            ], style={'display': 'flex', 'gap': '20px', 'marginBottom': '20px'}),

            html.Div([dcc.Graph(figure=unk_fig, style={'height': '300px'})], style=card_style),

            html.Div([
                html.H3("⚠️ Unknown Face Log",
                        style={'color': t['text'], 'marginBottom': '15px'}),
                dash_table.DataTable(
                    data=unknown_df.to_dict('records') if not unknown_df.empty else [],
                    columns=[{'name': c, 'id': c} for c in unknown_df.columns] if not unknown_df.empty else [],
                    filter_action='native',
                    sort_action='native',
                    page_size=10,
                    **table_style
                )
            ], style=card_style),
        ])

# ─────────────────────────────────────────
# EXPORT CALLBACK
# ─────────────────────────────────────────
@app.callback(
    Output('download-excel', 'data'),
    Input('export-btn', 'n_clicks'),
    State('tabs', 'value'),
    prevent_initial_call=True
)
def export_excel(n_clicks, tab):
    if not n_clicks:
        return None
    df    = load_all_attendance()
    today = datetime.now().strftime('%Y-%m-%d')
    if tab == 'today':
        data = df[df['Date'] == today]
    elif tab == 'weekly':
        week_start = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        data = df[df['Date'] >= week_start]
    elif tab == 'monthly':
        current_month = datetime.now().strftime('%Y-%m')
        data = df[df['Date'].str.startswith(current_month)]
    else:
        data = df
    if data.empty:
        return None
    path = export_to_excel(data, tab)
    return dcc.send_file(path)

# ─────────────────────────────────────────
# RUN
# ─────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 45)
    print("  📊 SMART ATTENDANCE DASHBOARD")
    print("=" * 45)
    print("\n🌐 Open your browser:")
    print("   http://127.0.0.1:8050")
    print("\n🔒 Login with your credentials")
    print("\nFeatures:")
    print("  🌙 Dark / ☀️  Light theme")
    print("  📅 Today / Weekly / Monthly tabs")
    print("  🗺️  Attendance Heatmap")
    print("  ⚠️  Unknown faces tracker")
    print("  📥 Export to Excel")
    print("\nPress Ctrl+C to stop\n")
    app.run(debug=True)
